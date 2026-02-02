"""SQLite-backed waybill storage using SQLModel."""

from __future__ import annotations

import json
import os
from contextlib import contextmanager
from pathlib import Path
from datetime import date, datetime
from typing import Any, Dict, Iterable, Optional, Tuple

from sqlmodel import Field, Session, SQLModel, create_engine, select

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = Path(os.getenv("WAYBILL_DB_PATH", PROJECT_ROOT / "data" / "waybills.db"))


class Waybill(SQLModel, table=True):
    id: Optional[int] = Field(default=None, primary_key=True) # ID
    waybill_no: str = Field(index=True, unique=True) # 运单号
    company: Optional[str] = None # 物流公司
    insured: Optional[bool] = None # 是否保价
    full_insured: Optional[bool] = None # 是否全额保价
    weight: Optional[float] = None # 重量
    signed: Optional[bool] = None # 是否签收
    signed_at: Optional[date] = None # 签收日期
    status: Optional[str] = None # 状态
    cost: Optional[float] = None # 费用
    price: Optional[float] = None # 物品价值
    route: Optional[str] = None  # 运输路线，JSON字符串


engine = create_engine(f"sqlite:///{DB_PATH}", echo=False)


def init_db() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    SQLModel.metadata.create_all(engine)


@contextmanager
def get_session():
    with Session(engine) as session:
        yield session


def _serialize_route(route: Any) -> Optional[str]:
    if route is None:
        return None
    if isinstance(route, str):
        return route
    try:
        return json.dumps(route, ensure_ascii=False)
    except TypeError:
        return None


def _deserialize_route(route: Optional[str]) -> Any:
    if not route:
        return None
    try:
        return json.loads(route)
    except json.JSONDecodeError:
        return route


def upsert_waybill(payload: Dict[str, Any]) -> Waybill:
    waybill_no = payload.get("waybill_no") or payload.get("waybillNo")
    if not waybill_no:
        raise ValueError("waybill_no is required")

    with get_session() as session:
        existing = session.exec(select(Waybill).where(Waybill.waybill_no == waybill_no)).first()
        data = payload.copy()
        data["waybill_no"] = waybill_no
        if "route" in data:
            data["route"] = _serialize_route(data["route"])
        if "signed_at" in data:
            data["signed_at"] = _normalize_date(data["signed_at"])

        if existing:
            for key, value in data.items():
                if hasattr(existing, key):
                    setattr(existing, key, value)
            session.add(existing)
            session.commit()
            session.refresh(existing)
            return existing

        record = Waybill(**{k: v for k, v in data.items() if hasattr(Waybill, k)})
        session.add(record)
        session.commit()
        session.refresh(record)
        return record


def query_waybill_db(waybill_no: str) -> Optional[Dict[str, Any]]:
    if not waybill_no:
        return None
    with get_session() as session:
        record = session.exec(select(Waybill).where(Waybill.waybill_no == waybill_no)).first()
        if not record:
            return None
        data = record.model_dump()
        data.pop("id", None)
        data["route"] = _deserialize_route(data.get("route"))
        return data


def bulk_upsert(payloads: Iterable[Dict[str, Any]]) -> int:
    count = 0
    for payload in payloads:
        upsert_waybill(payload)
        count += 1
    return count


def import_from_json(path: Path) -> int:
    if not path.exists():
        raise FileNotFoundError(f"{path} not found")
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    payloads = []
    if isinstance(data, dict):
        for waybill_no, payload in data.items():
            if isinstance(payload, dict):
                payload = payload.copy()
                payload["waybill_no"] = payload.get("waybill_no") or waybill_no
                payloads.append(payload)
    elif isinstance(data, list):
        payloads = [p for p in data if isinstance(p, dict)]
    return bulk_upsert(payloads)


def _normalize_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        text = str(value).strip()
        if not text:
            return None
        return date.fromisoformat(text)
    except Exception:
        return None


def _normalize_bool(value: Any) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"true", "yes", "y", "1", "是"}:
        return True
    if text in {"false", "no", "n", "0", "否"}:
        return False
    return None


def _normalize_route(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, (list, tuple)):
        return list(value)
    if isinstance(value, str) and "," in value:
        return [item.strip() for item in value.split(",") if item.strip()]
    return value


def import_from_excel_bytes(content: bytes) -> Tuple[int, int]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for Excel import")
    if not content:
        return 0, 0
    from io import BytesIO

    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return 0, 0

    header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h is not None}
    imported = 0
    skipped = 0

    def get(row: Tuple[Any, ...], key: str) -> Any:
        idx = header_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    for row in rows:
        waybill_no = get(row, "waybill_no")
        if not waybill_no:
            skipped += 1
            continue

        payload = {
            "waybill_no": str(waybill_no).strip(),
            "company": get(row, "company"),
            "insured": _normalize_bool(get(row, "insured")),
            "full_insured": _normalize_bool(get(row, "full_insured")),
            "weight": get(row, "weight"),
            "signed": _normalize_bool(get(row, "signed")),
            "signed_at": _normalize_date(get(row, "signed_at")),
            "status": get(row, "status"),
            "cost": get(row, "cost"),
            "price": get(row, "price"),
            "route": _normalize_route(get(row, "route")),
        }
        upsert_waybill(payload)
        imported += 1

    return imported, skipped
